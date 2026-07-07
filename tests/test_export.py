"""Tests for report exports."""

from openpyxl import load_workbook

from financial_planner.export import export_results_to_excel, export_results_to_markdown
from financial_planner.simulation import run_simulation
from test_simulation import sample_config


def test_markdown_export_contains_final_comparison(tmp_path) -> None:
    config = sample_config()
    results = run_simulation(config)

    path = export_results_to_markdown(results, config, tmp_path / "report.md")

    text = path.read_text(encoding="utf-8")
    assert "# Informe de planificación financiera" in text
    assert "Comparación final" in text
    assert "Ayudantes de decisión" in text
    assert "Resumen de escenarios" in text
    assert "Comprobación de coherencia" in text
    assert "Monte Carlo" in text
    assert "Resumen de sensibilidad" in text
    assert "solo_fondo_inversion" in text


def test_excel_export_contains_v1_sheets(tmp_path) -> None:
    config = sample_config()
    results = run_simulation(config)

    path = export_results_to_excel(results, config, tmp_path / "results.xlsx")
    workbook = load_workbook(path, read_only=True)

    assert {
        "resultados_anuales",
        "comparacion_final",
        "escenarios",
        "plantillas_escenario",
        "sensibilidad",
        "comprobacion",
        "monte_carlo",
        "productos",
        "decisiones",
        "avisos",
        "supuestos",
    }.issubset(set(workbook.sheetnames))
